# encoding: utf-8
from flask import Flask
import json
import random
from openpyxl import load_workbook
import sys
import pandas as pd

# set encoding
#reload(sys)
#sys.setdefaultencoding("utf-8")

# open data book
wb = load_workbook('data.xlsx', read_only=True)

# read regions list
ws = wb.get_sheet_by_name("Регионы")
rg_codes = [r[0].value for r in ws.iter_rows()]
rg_codes = rg_codes[1:]

# read macroregions
ws = wb.get_sheet_by_name("Макрорегионы")
mrg_codes = [r[0].value for r in ws.iter_rows()]
mrg_codes = mrg_codes[1:]

# Flask
app = Flask(__name__)

# returns color codes
def get_color(plan, fact):
	if abs(plan - fact) < 20:
		return 'Y'
	if plan > fact:
		return 'R'
	return 'G'

# create dummy object
LO = 50
HI = 150

def create_dummy():
	regions = []
	for i in rg_codes:
		plan = random.randint(LO, HI)
		fact = random.randint(LO, HI)
		regions.append(
				{
				'id': i,
				'plan': plan,
				'fact': fact,
				'color': get_color(plan, fact)
				}
			)
	mregions = []
	for i in mrg_codes:
		plan = random.randint(LO, HI)
		fact = random.randint(LO, HI)
		mregions.append(
				{
				'id': i,
				'plan': plan,
				'fact': fact,
				'color': get_color(plan, fact)
				}
			)
	return {
		'regions': regions,
		'macroregions': mregions
	}


dummy = create_dummy()

@app.route('/biserver/api/v1.0/regions', methods=['GET'])
def hello_world():
    return json.dumps(dummy)